import tkinter as tk
from pathlib import Path
from tkinter import ttk, messagebox
import requests
from PIL import Image, ImageTk
import io
import pyodbc
import threading
import queue
import asyncio
import json
import os
from dotenv import load_dotenv
import hashlib
import imagehash

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


from ShopScraper import octo_get_product_images, directdeal_get_product_images, api_get_product_images, easynotebooks_get_product_images, kosatec_get_product_images, dcs_get_product_images, incomgroup_get_product_images, wortmann_get_product_images
from ShopSelenium import fourcom_get_product_images, notebooksbilliger_get_product_images, komputronik_get_product_images, wave_get_product_images

DODAVATELE = {
    # klasickej scrape
    "octo it": {"kod": "348651", "produkt_dotaz_kod": "SivCode", "funkce": octo_get_product_images, "paralelně": True},
    "directdeal/everit": {"kod": "268493", "produkt_dotaz_kod": "SivCode", "funkce": directdeal_get_product_images, "paralelně": True},
    "api": {"kod": "161784", "produkt_dotaz_kod": "SivCode", "funkce": api_get_product_images, "paralelně": True},
    "NetFactory/easynotebooks": {"kod": "351191", "produkt_dotaz_kod": "SivCode", "funkce": easynotebooks_get_product_images, "paralelně": True},
    "Kosatec": {"kod": "165463", "produkt_dotaz_kod": "SivCode", "funkce": kosatec_get_product_images, "paralelně": True},
    "Dcs (nekvalitní)": {"kod": "319004", "produkt_dotaz_kod": "SivCode", "funkce": dcs_get_product_images, "paralelně": True},
    "IncomGroup": {"kod": "169701", "produkt_dotaz_kod": "SivCode2", "funkce": incomgroup_get_product_images, "paralelně": True},
    "Wortmann": {"kod": "190157", "produkt_dotaz_kod": "SivCode", "funkce": wortmann_get_product_images, "paralelně": True},


    # selenium
    "Wave (selenium)": {"kod": "115565", "produkt_dotaz_kod": "SivCode", "funkce": wave_get_product_images, "paralelně": False},
    "notebooksbilliger (selenium)": {"kod": "340871", "produkt_dotaz_kod": "SivCode", "funkce": notebooksbilliger_get_product_images, "paralelně": False},
    "fourcom (selenium)": {"kod": "312585", "produkt_dotaz_kod": "SivCode", "funkce": fourcom_get_product_images, "paralelně": False},
    "Komputronik (selenium)": {"kod": "104584", "produkt_dotaz_kod": "SivCode", "funkce": komputronik_get_product_images, "paralelně": False},
}

# Slovní volby počtu produktů pro combobox
POCTY_PRODUKTU = ["hodně málo", "málo", "středně", "hodně", "nejvíc"]

# Mapování slov -> čísel, se kterými pracuje logika (TOP ...)
POCET_MAP = {
    "hodně málo": 15,
    "málo": 25,
    "středně": 50,
    "hodně": 100,
    "nejvíc": 150,
}
OBRAZKY_NA_RADEK = ["2", "3", "4", "5", "6", "nekonečno"]

# Nová konstanta pro soubor s ignorovanými produkty
IGNORE_FILE = "ignoreSivCode.json"

# Nové konstanty pro práci s logování obrázků
EXCEL_LOG_PATH = "obrazky_log.xlsx"

# Přidáno: Konstanty pro práci s obrázky
IMG_DIR = "img"
SIMILARITY_THRESHOLD = 5  # Pro imagehash (max rozdíl 64, menší hodnota = větší podobnost)


class LoadingScreen:
    def __init__(self, root):
        self.root = root
        self.loading_window = tk.Toplevel(root)
        self.loading_window.title("Načítání...")
        self.loading_window.geometry("300x150")
        self.loading_window.resizable(False, False)

        # Center the loading window
        window_width = 300
        window_height = 150
        screen_width = self.loading_window.winfo_screenwidth()
        screen_height = self.loading_window.winfo_screenheight()
        position_top = int(screen_height / 2 - window_height / 2)
        position_right = int(screen_width / 2 - window_width / 2)
        self.loading_window.geometry(f"{window_width}x{window_height}+{position_right}+{position_top}")

        self.overlay = None  # Přidáno pro překryvnou obrazovku

        # Make it modal
        self.loading_window.grab_set()
        self.loading_window.transient(root)

        # Loading label
        tk.Label(self.loading_window, text="Načítám produkty...", font=("Arial", 14)).pack(pady=20)

        # Progress bar
        self.progress = ttk.Progressbar(
            self.loading_window,
            orient='horizontal',
            mode='indeterminate',
            length=200
        )
        self.progress.pack(pady=10)
        self.progress.start()

        # Disable close button
        self.loading_window.protocol("WM_DELETE_WINDOW", lambda: None)

    def close(self):
        self.progress.stop()
        self.loading_window.grab_release()
        self.loading_window.destroy()


class ObrFormApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Doplnění obrázků k produktům")
        self.root.geometry("1400x900")

        self.conn = None
        self.cursor = None
        self.filtrovane_produkty = []
        self.img_refs = {}
        self.vybrany_dodavatel = None
        self.vybrany_dodavatel_kod = None
        self.df = None
        self.produkty_k_zpracovani = []
        self.produkt_widgety = {}
        self.buffer_size = 25
        self.image_queue = queue.Queue()
        self.loading_threads = []
        self.all_check_var = tk.BooleanVar(value=True)  # zaškrtnutý checkbox pro "Vybrat vše"
        self.produkt_check_vars = {}
        self.image_check_vars = {}
        self.loading_active = False
        self.max_threads = 5
        self.image_cache = {}
        self.obrazky_na_radek = 6
        self.scrollregion_scheduled = False
        self.loading_screen = None

        # Nové atributy pro správu obrázků
        self.red_frame_images = set()  # {(kod, index)}
        self.original_images = {}  # {kod: [binární data obrázků]}
        self.existing_hashes = self.load_existing_image_hashes()  # Načte existující hashe obrázků

        # Načtení ignorovaných kódů při startu
        self.ignored_codes = self.load_ignored_codes()

        load_dotenv()
        db_table = os.getenv('DB_TABLE')

        self.featured_index = {}  # {kod_produktu: int | None} - index obrázku označeného jako #1
        self.image_frames = {}  # {kod_produktu: [frame_pro_jednotlivé_obrázky]}

        # Konfigurace databáze
        self.table_name = db_table
        self.column_mapping = {
            'code': 'SivCode',
            'name': 'SivName',
            'supplier': 'SivComId',
            'notes': 'SivNotePic',
            'pairing': 'SivStiId'
        }

        print("[DEBUG] Inicializace GUI...")
        self.setup_gui()

    def load_existing_image_hashes(self):
        """Načte hashe existujících obrázků v IMG_DIR"""
        hashes = set()
        if not os.path.exists(IMG_DIR):
            return hashes

        for root, _, files in os.walk(IMG_DIR):
            for file in files:
                if file.lower().endswith(('png', 'jpg', 'jpeg', 'gif', 'bmp')):
                    path = os.path.join(root, file)
                    try:
                        with Image.open(path) as img:
                            img_hash = imagehash.average_hash(img)
                            hashes.add(str(img_hash))
                    except Exception as e:
                        print(f"Chyba při načítání obrázku {path}: {e}")
        return hashes

    def is_image_similar(self, image_data):
        """Zjistí, zda je obrázek podobný nějakému existujícímu"""
        try:
            img = Image.open(io.BytesIO(image_data))
            new_hash = imagehash.average_hash(img)

            for existing_hash_str in self.existing_hashes:
                existing_hash = imagehash.hex_to_hash(existing_hash_str)
                if new_hash - existing_hash <= SIMILARITY_THRESHOLD:
                    return True
            return False
        except Exception as e:
            print(f"Chyba při porovnávání obrázků: {e}")
            return False

    def save_image_to_disk(self, image_data, supplier_code, siv_code, index):
        """Uloží obrázek na disk a přidá jeho hash"""
        try:
            supplier_dir = os.path.join(IMG_DIR, supplier_code)
            os.makedirs(supplier_dir, exist_ok=True)

            # Vytvořit unikátní název souboru
            file_hash = hashlib.md5(image_data).hexdigest()[:8]
            filename = f"{siv_code}_{index}_{file_hash}.jpg"
            path = os.path.join(supplier_dir, filename)

            with open(path, 'wb') as f:
                f.write(image_data)

            # Přidat hash do existujících
            img = Image.open(io.BytesIO(image_data))
            img_hash = imagehash.average_hash(img)
            self.existing_hashes.add(str(img_hash))
            return True
        except Exception as e:
            print(f"Chyba při ukládání obrázku: {e}")
            return False

    def mark_with_red_frame(self, kod, index):
        """Označí obrázek červeným rámečkem a odškrtne checkbox"""
        key = (kod, index)

        # Přepnout stav označení
        if key in self.red_frame_images:
            self.red_frame_images.remove(key)
            frame = self.image_frames[kod][index]
            frame.configure(highlightthickness=0)
        else:
            self.red_frame_images.add(key)
            frame = self.image_frames[kod][index]
            frame.configure(highlightbackground="red", highlightthickness=2)

            # Odškrtnout obrázek
            self.image_check_vars[kod][index].set(False)

    def load_ignored_codes(self):
        """Načte ignorované kódy z JSON souboru"""
        try:
            if os.path.exists(IGNORE_FILE):
                with open(IGNORE_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            return {}
        except Exception as e:
            print(f"[CHYBA] Načtení ignorovaných kódů: {e}")
            return {}

    def save_ignored_codes(self):
        """Uloží ignorované kódy do JSON souboru"""
        try:
            with open(IGNORE_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.ignored_codes, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"[CHYBA] Ukládání ignorovaných kódů: {e}")

    def add_ignored_code(self, supplier_code, siv_code):
        """Přidá kód do seznamu ignorovaných pro daného dodavatele"""
        if supplier_code not in self.ignored_codes:
            self.ignored_codes[supplier_code] = []

        if siv_code not in self.ignored_codes[supplier_code]:
            self.ignored_codes[supplier_code].append(siv_code)
            self.save_ignored_codes()

    def connect_to_database(self):
        """Připojí se k SQL Serveru"""
        try:
            # Load environment variables from .env file
            load_dotenv()

            server = os.getenv('DB_SERVER')
            database = os.getenv('DB_DATABASE')
            username = os.getenv('DB_USERNAME')
            password = os.getenv('DB_PASSWORD')

            conn_str = (
                f'DRIVER={{SQL Server}};'
                f'SERVER={server};'
                f'DATABASE={database};'
                f'UID={username};'
                f'PWD={password}'
            )

            print("[DEBUG] Pokus o připojení k databázi...")
            self.conn = pyodbc.connect(conn_str)
            self.cursor = self.conn.cursor()
            print("[DEBUG] Úspěšně připojeno k SQL Serveru")
            return True
        except Exception as e:
            print(f"[CHYBA] Připojení k databázi: {str(e)}")
            messagebox.showerror("Chyba", f"Chyba při připojování k databázi:\n{str(e)}")
            return False

    def check_database_structure(self):
        """Zkontroluje existenci potřebných sloupců"""
        try:
            required_columns = ['SivCode', 'SivComId', 'SivNotePic', 'SivName']

            query = f"""
                SELECT COLUMN_NAME 
                FROM INFORMATION_SCHEMA.COLUMNS 
                WHERE TABLE_NAME = '{self.table_name}'
            """
            self.cursor.execute(query)
            existing_columns = [row.COLUMN_NAME for row in self.cursor.fetchall()]

            missing_columns = [col for col in required_columns if col not in existing_columns]

            if missing_columns:
                messagebox.showerror("Chyba",
                                     f"V tabulce chybí potřebné sloupce!\n"
                                     f"Chybějící sloupce: {', '.join(missing_columns)}")
                return False

            print("[DEBUG] Všechny potřebné sloupce existují")
            return True

        except Exception as e:
            print(f"[CHYBA] Při kontrole struktury databáze: {str(e)}")
            messagebox.showerror("Chyba", f"Chyba při kontrole struktury databáze:\n{str(e)}")
            return False

    def setup_gui(self):
        """Vytvoří GUI prvky aplikace."""
        top_frame = tk.Frame(self.root)
        top_frame.pack(fill=tk.X, padx=10, pady=10)

        # Combobox pro výběr dodavatele
        tk.Label(top_frame, text="Dodavatel:", font=("Arial", 12)).pack(side=tk.LEFT, padx=5)
        self.combo_dodavatel = ttk.Combobox(top_frame, values=list(DODAVATELE.keys()), state="readonly",
                                            font=("Arial", 12), width=20)
        self.combo_dodavatel.pack(side=tk.LEFT, padx=5)
        self.combo_dodavatel.bind("<<ComboboxSelected>>", self.combo_selected)

        # Combobox pro výběr počtu produktů
        tk.Label(top_frame, text="Počet produktů:", font=("Arial", 12)).pack(side=tk.LEFT, padx=(20, 5))
        self.combo_pocet = ttk.Combobox(
            top_frame,
            values=POCTY_PRODUKTU,
            state="readonly",
            font=("Arial", 12),
            width=12  # delší texty se lépe vejdou
        )
        self.combo_pocet.bind("<<ComboboxSelected>>", self.update_buffer_size)
        self.combo_pocet.current(2)  # default: "středně"
        self.combo_pocet.pack(side=tk.LEFT, padx=5)
        # hned nastaví buffer_size podle defaultu
        self.update_buffer_size()

        # Combobox pro výběr počtu obrázků na řádek
        tk.Label(top_frame, text="Obrázky na řádek:", font=("Arial", 12)).pack(side=tk.LEFT, padx=(20, 5))
        self.combo_obrazky_na_radek = ttk.Combobox(top_frame, values=OBRAZKY_NA_RADEK, state="readonly",
                                                   font=("Arial", 12), width=10)
        self.combo_obrazky_na_radek.current(4)  # 4. pozice toho listu takze (6)
        self.combo_obrazky_na_radek.pack(side=tk.LEFT, padx=5)
        self.combo_obrazky_na_radek.bind("<<ComboboxSelected>>", self.update_obrazky_na_radek)

        # Checkbox "Vybrat vše"
        self.chk_all = tk.Checkbutton(top_frame, text="Vybrat vše", variable=self.all_check_var,
                                      font=("Arial", 12), command=self.toggle_all)
        self.chk_all.pack(side=tk.LEFT, padx=20)

        # Canvas s scrollbarem
        self.canvas_frame = tk.Frame(self.root)
        self.canvas_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.canvas = tk.Canvas(self.canvas_frame)
        self.scroll_y = tk.Scrollbar(self.canvas_frame, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.scroll_y.set)

        self.scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.inner_frame = tk.Frame(self.canvas)
        self.canvas.create_window((0, 0), window=self.inner_frame, anchor=tk.NW)

        self.inner_frame.bind("<Configure>", lambda e: self.schedule_scrollregion_update())
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.inner_frame.bind("<MouseWheel>", self._on_mousewheel)

        # PŘIDÁNO: Scrollování v celé aplikaci
        self.root.bind("<MouseWheel>", self._on_mousewheel)

        # Tlačítka dole
        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=10)

        tk.Button(btn_frame, text="Potvrdit", command=self.potvrdit_vse,
                  font=("Arial", 14), height=2, width=15).pack(side=tk.LEFT, padx=20)
        tk.Button(btn_frame, text="Zrušit", command=self.zrusit_vse,
                  font=("Arial", 14), height=2, width=15).pack(side=tk.LEFT, padx=20)

    def schedule_scrollregion_update(self):
        """Plánuje aktualizaci scrollregionu pro plynulejší scrollování."""
        if not self.scrollregion_scheduled:
            self.scrollregion_scheduled = True
            self.root.after(100, self.update_scrollregion)

    def update_buffer_size(self, event=None):
        """Aktualizuje buffer_size podle slovní volby v comboboxu Počet produktů."""
        try:
            volba = str(self.combo_pocet.get()).strip()
            # POCET_MAP je globální konstanta z horní části souboru
            self.buffer_size = POCET_MAP.get(volba, 50)  # fallback na 'středně' = 50
        except Exception:
            self.buffer_size = 50
        print(f"[DEBUG] buffer_size -> {self.buffer_size} (volba: {self.combo_pocet.get()})")

    def update_scrollregion(self):
        """Aktualizuje scrollregion canvasu s kontrolou existence."""
        if self.inner_frame.winfo_exists():
            self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        self.scrollregion_scheduled = False

    def update_obrazky_na_radek(self, event=None):
        """Aktualizuje počet obrázků na řádek podle výběru uživatele."""
        vyber = self.combo_obrazky_na_radek.get()
        if vyber == "nekonečno":
            self.obrazky_na_radek = float('inf')
        else:
            self.obrazky_na_radek = int(vyber)

        # Přerozdělení obrázků podle nového nastavení
        for kod, data in self.produkt_widgety.items():
            self.reorganize_images(data['images_frame'], data['urls'], kod)

    def reorganize_images(self, frame, urls, kod):
        """Přerozdělí obrázky v frame podle aktuálního nastavení počtu na řádek."""
        # Odstranit staré widgety
        for widget in frame.winfo_children():
            widget.destroy()

        # Znovu sestavíme registry frame-ů pro daný produkt
        self.image_frames[kod] = []

        current_col = 0
        row_frame = None

        for i, url in enumerate(urls):
            if current_col % (self.obrazky_na_radek if self.obrazky_na_radek != float('inf') else 999999) == 0:
                row_frame = tk.Frame(frame)
                row_frame.pack(fill=tk.X)
                current_col = 0

            # Frame pro jeden obrázek
            img_frame = tk.Frame(row_frame, bd=0, highlightthickness=0)
            img_frame.grid(row=0, column=current_col, padx=5, pady=5)
            current_col += 1
            self.image_frames[kod].append(img_frame)

            # Nastavení červeného rámečku pokud je obrázek označen
            if (kod, i) in self.red_frame_images:
                img_frame.configure(highlightbackground="red", highlightthickness=2)

            # Checkbox
            img_var = self.image_check_vars[kod][i]
            chk = tk.Checkbutton(
                img_frame,
                variable=img_var,
                command=lambda k=kod: self.update_product_check(k)
            )
            chk.pack()

            # Label s obrázkem
            label = tk.Label(img_frame, image=self.img_refs[kod][i])
            label.image = self.img_refs[kod][i]
            label.pack()

            # Bind událostí
            label.bind("<Button-1>", lambda e, var=img_var: var.set(not var.get()))
            label.bind("<Button-2>", lambda e, k=kod, idx=i: self.mark_with_red_frame(k, idx))
            label.bind("<Button-3>", lambda e, k=kod, idx=i: self.mark_featured(k, idx))

            # Overlay "1" pro featured obrázek
            overlay = tk.Label(img_frame, text="1", bg="white", fg="black", font=("Arial", 14, "bold"))
            overlay._is_overlay = True
            if self.featured_index.get(kod) == i:
                overlay.place(x=5, y=5)
                img_frame.configure(highlightbackground="black", highlightthickness=2)
            else:
                overlay.place_forget()

    def _on_mousewheel(self, event):
        """Zpracování scrollování myší s akcelerací."""
        scroll_amount = int(-1 * (event.delta / 40))
        self.canvas.yview_scroll(scroll_amount, "units")

    def combo_selected(self, event):
        """Zpracuje výběr dodavatele a počtu produktů (při každém reloadu vyjet nahoru)."""
        self.vybrany_dodavatel = self.combo_dodavatel.get()
        info = DODAVATELE[self.vybrany_dodavatel]
        self.vybrany_dodavatel_kod = info["kod"]
        self.vybrana_funkce = info["funkce"]

        print(
            f"[DEBUG] Vybrán dodavatel: {self.vybrany_dodavatel}..., kód: {self.vybrany_dodavatel_kod}, počet: {self.buffer_size}")

        # Zobrazit černou překryvnou obrazovku
        self.show_overlay()

        # Vytvořit loading screen (bude nad černou obrazovkou)
        self.loading_screen = LoadingScreen(self.root)
        self.root.update()  # Force update to show loading screen immediately

        # <<< přidej: okamžitě vyjet nahoru, aby nová várka začínala vždy "odshora"
        try:
            self.canvas.yview_moveto(0.0)
        except Exception as e:
            print(f"[WARN] combo_selected: scroll nahoru selhal: {e}")

        # Zakázat UI prvky během načítání
        self.combo_dodavatel.config(state='disabled')
        self.combo_pocet.config(state='disabled')
        self.combo_obrazky_na_radek.config(state='disabled')
        self.chk_all.config(state='disabled')

        # Spustit načítání v samostatném vlákně
        loading_thread = threading.Thread(target=self.load_products_thread, daemon=True)
        loading_thread.start()

    def show_overlay(self):
        """Zobrazí černou překryvnou obrazovku a deaktivuje UI"""
        self.overlay = tk.Canvas(self.root, bg='black', highlightthickness=0)
        self.overlay.place(x=0, y=0, relwidth=1, relheight=1)

        # Opraveno: Pro Canvas používáme jiný způsob pro zvednutí nad ostatní widgety
        self.overlay.tag_raise(tk.ALL)  # Místo self.overlay.lift()

        # Přidat průhlednost (volitelné)
        try:
            self.overlay.attributes('-alpha', 0.7)
        except:
            pass  # Některé platformy nepodporují průhlednost

    def hide_overlay(self):
        """Skryje černou překryvnou obrazovku"""
        if hasattr(self, 'overlay') and self.overlay:
            self.overlay.destroy()
            self.overlay = None

    def load_products_thread(self):
        """Thread for loading products to keep UI responsive (VERBOSE DB READ)."""
        import pandas as pd
        try:
            # Připojení k DB
            if not self.connect_to_database():
                self.root.after(0, self.loading_screen.close)
                return

            if not self.check_database_structure():
                self.close_database()
                self.root.after(0, self.loading_screen.close)
                return

            # Získání informací o dodavateli
            supplier_info = DODAVATELE[self.vybrany_dodavatel]
            self.vybrany_dodavatel_kod = supplier_info["kod"]
            self.vybrana_funkce = supplier_info["funkce"]
            produkt_dotaz_kod = supplier_info["produkt_dotaz_kod"]

            # Ignorované kódy
            ignored_codes = self.ignored_codes.get(self.vybrany_dodavatel_kod, [])

            # Přípravná tabulka
            self.cursor.execute("""
                CREATE TABLE #IgnoredCodes (
                    SivCode VARCHAR(50) COLLATE DATABASE_DEFAULT
                )
            """)

            if ignored_codes:
                self.cursor.executemany("INSERT INTO #IgnoredCodes VALUES (?)",
                                        [(code,) for code in ignored_codes])

            # SQL dotaz - přidáno SivCode2
            query = f"""
                SELECT TOP {self.buffer_size} 
                    [{produkt_dotaz_kod}] AS SivCode, 
                    SivName,
                    SivCode2
                FROM [{self.table_name}]
                join StoItem with(nolock) on (SivStiId = StiId)
                left join SCategory with(nolock) on (StiScaId = ScaId)
                WHERE [{self.column_mapping['supplier']}] = ?
                  AND SivOrdVen = 1
                  AND not exists (Select top 1 1 from Attach with(nolock) where AttSrcId = StiId and AttPedId = 52 and (AttTag like 'sys-gal%' or AttTag = 'sys-enl' or AttTag = 'sys-thu'))
                  AND StiPLPict is null
                  AND ScaId not in (8843,8388,8553,8387,6263,8231,7575,5203,2830,269,1668,2391,1634,7209)
                  AND ([{self.column_mapping['notes']}] IS NULL OR [{self.column_mapping['notes']}] = '')
                  AND ([{self.column_mapping['pairing']}] IS NOT NULL AND [{self.column_mapping['pairing']}] <> '')
                  AND StiHide = 0
                  AND StiHideI = 0
                  AND NOT EXISTS (
                      SELECT 1 FROM #IgnoredCodes 
                      WHERE SivCode = [{self.table_name}].[{produkt_dotaz_kod}] COLLATE DATABASE_DEFAULT
                  ) ORDER BY NEWID()
            """
            params = [self.vybrany_dodavatel_kod]

            # --- VERBOSE VÝPIS PŘED DOTAZEM ---
            print("\n" + "=" * 80)
            print("[DB READ] START – výběr produktů bez obrázků")
            print("-" * 80)
            print("Dodavatel:", self.vybrany_dodavatel, "| Kód:", self.vybrany_dodavatel_kod)
            print("Tabulka:", self.table_name)
            print("TOP:", self.buffer_size)
            print("Parametry:", params)
            print("SQL:")
            print(query.strip())
            print("-" * 80)

            # Provedení dotazu a převod do tabulky
            self.cursor.execute(query, params)
            rows = self.cursor.fetchall()
            self.filtrovane_produkty = [{'SivCode': r.SivCode, 'SivName': r.SivName, 'SivCode2': r.SivCode2} for r in rows]

            df = pd.DataFrame(self.filtrovane_produkty, columns=["SivCode", "SivName", "SivCode2"])
            print(f"Počet načtených řádků: {len(df)} (ignorováno: {len(ignored_codes)})")
            if not df.empty:
                # Přehledná tabulka (bez indexu)
                print("\nNÁHLED DAT (max 50 řádků):")
                print(df.head(50).to_string(index=False))
            else:
                print("\nVýsledek je prázdný.")

            print("=" * 80 + "\n")
            # --- /VERBOSE ---

            # Uzavření DB (jen pro čtení)
            self.close_database()

            if not self.filtrovane_produkty:
                self.root.after(0, lambda: messagebox.showinfo("Info", "Žádné produkty k doplnění."))
                self.root.after(0, self.hide_overlay)
                self.root.after(0, self.loading_screen.close)
                return

            # Vyčištění GUI
            self.root.after(0, self.clear_gui)

            self.produkty_k_zpracovani = self.filtrovane_produkty[:]
            self.produkt_widgety = {}
            self.produkt_check_vars = {}
            self.image_check_vars = {}
            self.img_refs = {}
            self.all_check_var.set(False)

            # Spuštění načítání obrázků
            self.start_async_image_loading()

        except Exception as e:
            print(f"[CHYBA] Při načítání produktů: {e}")
            self.root.after(0, lambda: messagebox.showerror("Chyba", f"Chyba při načítání produktů:\n{e}"))
            self.root.after(0, self.loading_screen.close)
            self.root.after(0, self.hide_overlay)
            self.close_database()

    def mark_featured(self, kod, index):
        """Označí/vypne 'první' obrázek u produktu (pravý klik)."""
        current = self.featured_index.get(kod)
        # Toggle: klik na stejný index zruší výběr
        if current == index:
            self.featured_index[kod] = None
        else:
            self.featured_index[kod] = index

        # Vizualizace: ukázat/skryt overlay '1' a zvýraznit rámeček
        frames = self.image_frames.get(kod, [])
        for i, f in enumerate(frames):
            # hledej overlay label vytvořený v reorganize_images (má příznak _is_overlay = True)
            overlay = None
            for child in f.winfo_children():
                if isinstance(child, tk.Label) and getattr(child, "_is_overlay", False):
                    overlay = child
                    break
            if overlay:
                if self.featured_index[kod] == i:
                    overlay.place(x=5, y=5)  # zobrazíme placku "1" vlevo nahoře
                    f.configure(highlightbackground="black", highlightthickness=2)
                else:
                    overlay.place_forget()
                    f.configure(highlightthickness=0)

    def clear_gui(self):
        """Clear the GUI in the main thread + posunout scroll nahoru."""
        # Vyčistit původní obrázky
        self.original_images = {}
        self.red_frame_images.clear()

        # Odstranit všechny widgety z inner_frame
        for widget in self.inner_frame.winfo_children():
            widget.destroy()

        # Resetovat stav
        self.produkt_widgety = {}
        self.produkt_check_vars = {}
        self.image_check_vars = {}
        self.img_refs = {}
        self.featured_index = {}
        self.image_frames = {}
        self.all_check_var.set(False)

        # Aktualizovat GUI a VŽDY srolovat nahoru
        try:
            # nejdřív refresh, pak skok na začátek
            self.canvas.update_idletasks()
            self.canvas.yview_moveto(0.0)  # <<<<<< klíčová řádka
        except Exception as e:
            print(f"[WARN] clear_gui: scroll nahoru se nepodařil: {e}")

        # pro jistotu ještě přepočítat scrollregion (nebude vadit ani když je prázdný)
        try:
            self.schedule_scrollregion_update()
        except Exception:
            pass

    def start_async_image_loading(self):
        """Spustí asynchronní načítání obrázků s optimalizovaným počtem vláken."""
        if not self.loading_active:
            self.loading_active = True

            # Získání informací o paralelním režimu pro dodavatele
            supplier_info = DODAVATELE[self.vybrany_dodavatel]
            paralelne = supplier_info.get("paralelně", True)  # Výchozí True pokud není definováno

            # Nastavení max_threads podle paralelního režimu
            if paralelne:
                max_threads = self.max_threads
            else:
                max_threads = 1  # Sekvenční režim

            print(f"[THREAD] Spouštím {max_threads} vláken (paralelně: {paralelne})")

            for _ in range(min(max_threads, len(self.produkty_k_zpracovani))):
                if self.produkty_k_zpracovani:
                    produkt = self.produkty_k_zpracovani.pop(0)
                    t = threading.Thread(target=self.load_product_images, args=(produkt,))
                    t.daemon = True
                    t.start()
                    self.loading_threads.append(t)

            # Pravidelně kontrolovat stav
            self.root.after(500, self.check_threads)

    def check_threads(self):
        """Kontroluje stav načítacích vláken a spouští nová podle potřeby."""
        # Odstranění ukončených vláken
        self.loading_threads = [t for t in self.loading_threads if t.is_alive()]

        # Získání informací o paralelním režimu pro dodavatele
        supplier_info = DODAVATELE[self.vybrany_dodavatel]
        paralelne = supplier_info.get("paralelně", True)  # Výchozí True pokud není definováno

        # Nastavení max_threads podle paralelního režimu
        if paralelne:
            max_threads = self.max_threads
        else:
            max_threads = 1  # Sekvenční režim

        free_slots = max_threads - len(self.loading_threads)

        # Spuštění nových vláken pro volné sloty
        for _ in range(min(free_slots, len(self.produkty_k_zpracovani))):
            if self.produkty_k_zpracovani:
                produkt = self.produkty_k_zpracovani.pop(0)
                t = threading.Thread(target=self.load_product_images, args=(produkt,))
                t.daemon = True
                t.start()
                self.loading_threads.append(t)

        # Kontrola dokončení
        if not self.loading_threads and not self.produkty_k_zpracovani:
            self.loading_active = False
            print("[THREAD] Všechna vlákna dokončena")
            self.root.after(0, self.enable_ui_elements)
            self.root.after(0, self.loading_screen.close)
            self.root.after(0, self.hide_overlay)
        else:
            # Plánování další kontroly
            self.root.after(200, self.check_threads)

    def enable_ui_elements(self):
        """Re-enable UI elements after loading is complete"""
        self.combo_dodavatel.config(state='readonly')
        self.combo_pocet.config(state='readonly')
        self.combo_obrazky_na_radek.config(state='readonly')
        self.chk_all.config(state='normal')
        self.chk_all.select()  # Select all by default

    def load_product_images(self, produkt):
        """Načte obrázky pro daný produkt ve worker vlákně.
        Nevykresluje nic předem – UI se vytvoří až při prvním úspěšném obrázku.
        """
        try:
            kod = produkt['SivCode']
            print(f"[THREAD] Načítám obrázky pro produkt: {kod}")

            # Příprava úložiště pro originální obrázky (naplní se až v add_single_image)
            if kod not in self.original_images:
                self.original_images[kod] = []

            # 1) Získání URL obrázků z funkce pro zvoleného dodavatele (většinou async)
            funkce_pro_dodavatele = self.vybrana_funkce
            if not funkce_pro_dodavatele:
                print(f"[CHYBA] Pro dodavatele {self.vybrany_dodavatel_kod} není definována funkce")
                return

            urls = []
            if asyncio.iscoroutinefunction(funkce_pro_dodavatele):
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    urls = loop.run_until_complete(funkce_pro_dodavatele(produkt['SivCode']))
                finally:
                    loop.close()
            else:
                urls = funkce_pro_dodavatele(produkt['SivCode'])

            # Normalizace návratu (povolíme tuple/dict, ale výsledkem má být list URL)
            if isinstance(urls, dict) and 'urls' in urls:
                urls = urls['urls']
            if not isinstance(urls, (list, tuple)):
                urls = []

            # 2) Stažení obrázků (bytes) a základní filtry
            valid_pairs = []  # (url, image_data)
            for url in urls:
                if not url:
                    continue
                try:
                    resp = requests.get(url, timeout=10)
                    if resp.status_code != 200:
                        continue
                    content_type = resp.headers.get("Content-Type", "")
                    if "image" not in content_type.lower():
                        continue
                    image_data = resp.content
                    # Volitelná deduplikace pomocí hashů (pokud je metoda k dispozici)
                    try:
                        if hasattr(self, "is_image_similar") and self.is_image_similar(image_data):
                            # podobný existujícím – přeskočit
                            continue
                    except Exception as e:
                        print(f"[WARN] Deduplikace selhala: {e}")

                    valid_pairs.append((url, image_data))
                except Exception as e:
                    print(f"[INFO] Stažení obrázku selhalo ({url}): {e}")

            if not valid_pairs:
                # Žádný použitelný obrázek -> nic nevykreslovat (žádný prázdný groupbox)
                print(f"[INFO] {kod}: Nenačten žádný použitelný obrázek.")
                return

            # 3) Předání do UI threadu – vytvoření PhotoImage až v add_single_image
            for url, image_data in valid_pairs:
                try:
                    self.root.after_idle(self.add_single_image, produkt, url, image_data)
                except Exception as e:
                    print(f"[INFO] Chyba při plánování add_single_image pro {url}: {e}")

        except Exception as e:
            print(f"[CHYBA] Při načítání obrázků: {e}")

    def display_product_with_images(self, produkt):
        """Zobrazí základní informace o produktu."""
        kod = produkt['SivCode']
        nazev = produkt.get('SivName', "")
        kod2 = produkt.get('SivCode2', "")

        # Frame pro celý produkt - TEXT JE NYNÍ PRÁZDNÝ
        frame_produkt = tk.LabelFrame(
            self.inner_frame,
            text="",  # Prázdný text, nahradíme vlastním widgetem
            font=("Arial", 12, "bold"),
            padx=10,
            pady=10,
        )
        frame_produkt.pack(fill=tk.X, padx=10, pady=5, ipadx=5, ipady=5)
        frame_produkt.grid_columnconfigure(0, weight=1)

        # Vytvoření klikatelného labelu s textem pro kopírování
        label_text = f"{kod} - {nazev} - {kod2}"
        label_nazev = tk.Label(
            frame_produkt,
            text=label_text,
            font=("Arial", 12, "bold"),
            cursor="hand2",  # Kurzor ruky při najetí
            fg="blue",  # Modrý text pro indikaci klikatelnosti
        )
        label_nazev.grid(row=0, column=0, sticky=tk.W, pady=(0, 10))

        # Bindování události pro kopírování textu
        label_nazev.bind("<Button-1>", lambda e: self.copy_to_clipboard(label_text))

        # Zbytek kódu zůstává stejný...
        # Checkbox pro výběr všech obrázků v produktu
        var_produkt = tk.BooleanVar(value=True)
        self.produkt_check_vars[kod] = var_produkt

        chk_produkt = tk.Checkbutton(
            frame_produkt,
            text="Vybrat všechny obrázky",
            variable=var_produkt,
            font=("Arial", 14, "bold"),
            command=lambda k=kod: self.toggle_product_images(k)
        )
        chk_produkt.grid(row=1, column=0, sticky=tk.W, pady=(0, 10))  # Pozor: změna row na 1

        # Frame pro obrázky
        frame_obrazky = tk.Frame(frame_produkt)
        frame_obrazky.grid(row=2, column=0, sticky=tk.W)  # Pozor: změna row na 2

        # Uložení widgetů
        self.produkt_widgety[kod] = {
            'frame': frame_produkt,
            'images_frame': frame_obrazky,
            'image_vars': [],
            'urls': [],
            'produkt': produkt
        }
        self.image_check_vars[kod] = []

        # Aktualizovat GUI
        self.canvas.update_idletasks()

    # Nová metoda pro kopírování do schránky
    def copy_to_clipboard(self, text):
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        # Volitelně zobrazit potvrzení
        print(f"Text zkopírován: {text}")

    def ignore_product(self, kod):
        """Přidá produkt do seznamu ignorovaných."""
        self.add_ignored_code(self.vybrany_dodavatel_kod, kod)
        if kod in self.produkt_widgety:
            self.produkt_widgety[kod]['frame'].destroy()
            del self.produkt_widgety[kod]
        messagebox.showinfo("Info", f"Produkt {kod} byl přidán do ignorovaných")

    def add_single_image(self, produkt, url, image_data):
        """BĚŽÍ V HLAVNÍM TK VLÁKNĚ.
        Bezpečně vytvoří PhotoImage; teprve při úspěchu zajistí vytvoření produktového frame.
        Pokud konverze selže, neudělá se nic (ani rám).
        """
        kod = produkt['SivCode']

        # 1) Pokus o vytvoření PhotoImage v HLAVNÍM vlákně
        try:
            img = Image.open(io.BytesIO(image_data))
            img.verify()  # základní kontrola validity (zničí file pointer)
            # Po verify je potřeba znovu otevřít
            img = Image.open(io.BytesIO(image_data))
            img.thumbnail((300, 300))
            photo = ImageTk.PhotoImage(img)
        except Exception as e:
            print(f"[INFO] add_single_image: selhalo vytvoření PhotoImage pro {kod}: {e}")
            return  # NIC nevykreslovat → žádný prázdný groupbox

        # 2) Pokud rám pro produkt ještě neexistuje, vytvoř ho teď (až po úspěšném obrázku)
        if kod not in self.produkt_widgety:
            self.display_product_with_images(produkt)

        # Ochranné inicializace struktur
        if kod not in self.featured_index:
            self.featured_index[kod] = None
        if kod not in self.image_frames:
            self.image_frames[kod] = []
        if kod not in self.img_refs:
            self.img_refs[kod] = []
        if kod not in self.image_check_vars:
            self.image_check_vars[kod] = []
        if kod not in self.original_images:
            self.original_images[kod] = []

        # 3) Uložení dat v JEDNOTNÉM pořadí (musí sedět indexy napříč vším)
        self.img_refs[kod].append(photo)  # reference na PhotoImage, aby je GC nesebral
        self.produkt_widgety[kod]['urls'].append(url)  # logická posloupnost URL
        self.original_images[kod].append(image_data)  # originální bytes (pro uložení / export)

        # 4) Checkbox proměnná (defaultně zaškrtnutý)
        img_var = tk.BooleanVar(value=True)
        self.image_check_vars[kod].append(img_var)
        self.produkt_widgety[kod]['image_vars'].append(img_var)

        # 5) Rozmístění podle nastavení + refresh
        self.reorganize_images(
            self.produkt_widgety[kod]['images_frame'],
            self.produkt_widgety[kod]['urls'],
            kod
        )

        self.root.update_idletasks()
        self.schedule_scrollregion_update()

    def toggle_all(self):
        """Vybere nebo zruší výběr všech obrázků u všech produktů."""
        select = self.all_check_var.get()
        for kod in self.produkt_check_vars:
            self.produkt_check_vars[kod].set(select)
            self.toggle_product_images(kod, select)

    def toggle_product_images(self, kod, value=None):
        """Vybere nebo zruší výběr všech obrázků v produktu."""
        if value is None:
            value = self.produkt_check_vars[kod].get()

        for var in self.image_check_vars[kod]:
            var.set(value)

    def update_product_check(self, kod):
        """Aktualizuje stav checkboxu produktu na základě obrázků."""
        # Zkontrolovat, zda stále existují image_vars
        if kod not in self.image_check_vars or not self.image_check_vars[kod]:
            return

        all_checked = all(var.get() for var in self.image_check_vars[kod])
        any_checked = any(var.get() for var in self.image_check_vars[kod])

        if all_checked:
            self.produkt_check_vars[kod].set(True)
        elif any_checked:
            # Pro částečný výběr necháme checkbox v "částečném" stavu
            pass
        else:
            self.produkt_check_vars[kod].set(False)

    def potvrdit_vse(self):
        """Uloží vybrané obrázky do DB + log (VERBOSE DB WRITE) a zobrazí přehlednou tabulku změn."""
        import pandas as pd

        if not self.connect_to_database():
            return

        excel_path = getattr(self, "EXCEL_LOG_PATH", "obrazky_log.xlsx")

        # Připrava Excel sešitu (beze změn v logice)
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
        except Exception:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log"

        if ws.max_row < 1 or ws["A1"].value is None or ws["B1"].value is None:
            ws["A1"].value = "ZAPSAL"
            ws["B1"].value = "NEPOTVRDIL"
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            bold_font = Font(bold=True)
            for cell in ("A1",):
                ws[cell].fill = green_fill
                ws[cell].font = bold_font
                ws[cell].alignment = Alignment(horizontal="center", vertical="center")
            for cell in ("B1",):
                ws[cell].fill = red_fill
                ws[cell].font = bold_font
                ws[cell].alignment = Alignment(horizontal="center", vertical="center")

        target_width_chars = self.px_to_excel_col_width(180)
        ws.column_dimensions["A"].width = target_width_chars
        ws.column_dimensions["B"].width = target_width_chars

        products_to_remove = []

        # --- sběr informací pro přehlednou tabulku zápisu ---
        pending_updates = []  # pro tabulku „co se bude zapisovat“
        skipped_or_ignored = []  # pro informaci co se neuložilo

        try:
            for kod, data in self.produkt_widgety.items():
                vybrane_urls = [
                    url for i, url in enumerate(data["urls"])
                    if i < len(data["image_vars"]) and data["image_vars"][i].get()
                ]

                fi = self.featured_index.get(kod)
                if fi is not None and 0 <= fi < len(data["urls"]):
                    featured_url = data["urls"][fi]
                    if featured_url in vybrane_urls:
                        vybrane_urls = [featured_url] + [u for u in vybrane_urls if u != featured_url]

                nevybrane_urls = [
                    url for i, url in enumerate(data["urls"])
                    if not (i < len(data["image_vars"]) and data["image_vars"][i].get())
                ]

                # Excel log (beze změn)
                next_row = ws.max_row + 1
                for url in vybrane_urls:
                    ws.cell(row=next_row, column=1).value = f'=OBRÁZEK("{url}","zapsal",1)'
                    ws.row_dimensions[next_row].height = self.px_to_points(100)
                    next_row += 1
                for url in nevybrane_urls:
                    ws.cell(row=next_row, column=2).value = f'=OBRÁZEK("{url}","nepotvrdil",1)'
                    ws.row_dimensions[next_row].height = self.px_to_points(100)
                    next_row += 1

                if vybrane_urls:
                    produkt = data["produkt"]
                    zapis = ";\n".join(vybrane_urls) + ";"

                    # --- VERBOSE buffer pro tabulku ---
                    pending_updates.append({
                        "SivCode": produkt["SivCode"],
                        "PocetURL": len(vybrane_urls),
                        "PrvniURL": vybrane_urls[0] if vybrane_urls else ""
                    })

                    # VLASTNÍ UPDATE
                    query = f"""
                        UPDATE [{self.table_name}]
                        SET [{self.column_mapping['notes']}] = ?
                        WHERE [{self.column_mapping['code']}] = ?
                    """
                    self.cursor.execute(query, (zapis, produkt["SivCode"]))
                    products_to_remove.append(kod)
                else:
                    self.add_ignored_code(self.vybrany_dodavatel_kod, kod)
                    products_to_remove.append(kod)
                    skipped_or_ignored.append({"SivCode": kod, "Důvod": "nic nevybráno → ignorováno"})

            # Uložení červeně označených na disk (beze změn)
            for (kod, index) in self.red_frame_images:
                if (kod in self.original_images and index < len(self.original_images[kod])):
                    image_data = self.original_images[kod][index]
                    self.save_image_to_disk(image_data, self.vybrany_dodavatel_kod, kod, index)

            # --- VERBOSE: shrnutí před commitem ---
            print("\n" + "=" * 80)
            print("[DB WRITE] START – zápis vybraných obrázků do DB")
            print("-" * 80)
            print("Tabulka:", self.table_name)
            if pending_updates:
                dfw = pd.DataFrame(pending_updates, columns=["SivCode", "PocetURL", "PrvniURL"])
                print("Plánované UPDATE (počet řádků):", len(dfw))
                print(dfw.to_string(index=False, max_colwidth=120))
            else:
                print("Nebude se zapisovat žádná položka (vše ignorováno / nic nevybráno).")

            if skipped_or_ignored:
                dfs = pd.DataFrame(skipped_or_ignored)
                print("\nPřeskočené/ignorované položky:")
                print(dfs.to_string(index=False))
            print("-" * 80)

            # Commit DB + uložení Excelu
            self.conn.commit()
            wb.save(excel_path)

            print("[DB WRITE] COMMIT OK")
            print("=" * 80 + "\n")
            # --- /VERBOSE ---

            # Post-processing Excel (tvoje stávající logika beze změn)
            com_result = {"status": "skipped", "reason": "missing_pywin32_or_excel"}
            try:
                try:
                    import win32com.client as win32  # pywin32
                except ImportError:
                    win32 = None

                if win32 is not None:
                    XL_CELL_TYPE_FORMULAS = -4123
                    excel = win32.Dispatch("Excel.Application")
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    excel.AskToUpdateLinks = False
                    excel.EnableEvents = False

                    wb_com = excel.Workbooks.Open(str(Path(excel_path).resolve()))
                    fixed = normalized = touched = skipped = 0
                    try:
                        for ws_com in wb_com.Worksheets:
                            try:
                                f_cells = ws_com.UsedRange.SpecialCells(XL_CELL_TYPE_FORMULAS)
                            except Exception:
                                continue
                            for cell in f_cells:
                                try:
                                    fl = cell.FormulaLocal
                                    if isinstance(fl, str) and fl.startswith("=@"):
                                        cell.Formula2Local = "=" + fl[2:]
                                        fixed += 1
                                    else:
                                        try:
                                            cell.Formula2Local = cell.Formula2Local
                                            normalized += 1
                                        except Exception:
                                            skipped += 1
                                    touched += 1
                                except Exception:
                                    skipped += 1
                                    continue
                        wb_com.Save()
                        com_result = {"status": "ok", "fixed": fixed, "normalized": normalized,
                                      "touched": touched, "skipped": skipped}
                    finally:
                        wb_com.Close(SaveChanges=False)
                        excel.Quit()
            except Exception as e:
                com_result = {"status": "error", "error": str(e)}

            print(f"[INFO] Excel COM normalization: {com_result}")

            if com_result.get("status") != "ok":
                try:
                    fixed = self.strip_at_from_formulas(excel_path)
                    print(f"[INFO] Fallback openpyxl: opraveno {fixed} buněk '=@'.")
                    if fixed == 0:
                        self.write_repair_instructions(excel_path)
                except Exception as e:
                    print(f"[WARN] Fallback openpyxl selhal: {e}")
                    self.write_repair_instructions(excel_path)

            #messagebox.showinfo("Info", "Všechny vybrané produkty byly uloženy.")

            for kod in products_to_remove:
                if kod in self.produkt_widgety:
                    self.produkt_widgety[kod]['frame'].destroy()
                    del self.produkt_widgety[kod]
                if kod in self.img_refs:
                    del self.img_refs[kod]

        except Exception as e:
            print(f"[CHYBA] Při ukládání do DB/Excelu: {e}")
            messagebox.showerror("Chyba", f"Chyba při ukládání:\n{e}")
        finally:
            self.close_database()
            self.red_frame_images.clear()
            if self.vybrany_dodavatel and self.vybrany_dodavatel_kod:
                self.combo_selected(None)
            else:
                self.hide_overlay()

    def zrusit_vse(self):
        """Zruší všechny produkty bez uložení."""
        for kod in list(self.produkt_widgety.keys()):
            self.produkt_widgety[kod]['frame'].destroy()
            del self.produkt_widgety[kod]
            if kod in self.img_refs:
                del self.img_refs[kod]
        self.hide_overlay()  # Přidáno skrytí overlay
        messagebox.showinfo("Info", "Všechny produkty byly zrušeny.")

    def close_database(self):
        """Uzavře databázové připojení."""
        if self.conn:
            try:
                self.conn.close()
                print("[DEBUG] Databázové připojení uzavřeno.")
            except:
                pass
            finally:
                self.conn = None
                self.cursor = None

    # --- Excel helpery ---

    def strip_at_from_formulas(self, path) -> int:
        """Odstraní úvodní '@' z každého vzorce ve všech listech. Vrátí počet oprav."""
        wb = load_workbook(path)
        fixed = 0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=False):
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.startswith("=@"):
                        cell.value = "=" + val[2:]
                        fixed += 1
        wb.save(path)
        return fixed

    def write_repair_instructions(self, path):
        """Zapíše do C2:D5 stručné instrukce k ruční opravě (Ctrl+H)."""
        wb = load_workbook(path)
        ws = wb.active
        ws["C2"] = "Instrukce k opravě:"
        ws["D2"] = "Zmáčknout Ctrl+H"
        ws["C3"] = "Najít:"
        ws["D3"] = "=@"
        ws["C4"] = "Nahradit:"
        ws["D4"] = "="
        ws["C5"] = "Nahradit vše"
        wb.save(path)

    def px_to_excel_col_width(self, pixels: int) -> float:
        # Excel šířka ~ (px - 5) / 7
        return max(0.0, (pixels - 5) / 7.0)

    def px_to_points(self, pixels: int) -> float:
        # 96 DPI -> 1 px = 0.75 pt
        return pixels * 0.75


if __name__ == "__main__":
    print("[DEBUG] Spouštím aplikaci...")
    root = tk.Tk()
    app = ObrFormApp(root)
    root.mainloop()